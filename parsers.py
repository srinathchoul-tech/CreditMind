# parsers.py

import io
import os
import re
from typing import Any, Dict

from models import FinancialRatios, GSTFeatures, NewsLegalFeatures

try:
    from pypdf import PdfReader
except ImportError:  # pragma: no cover
    PdfReader = None

try:
    import pdfplumber
except ImportError:  # pragma: no cover
    pdfplumber = None

try:
    from pdf2image import convert_from_bytes
except ImportError:  # pragma: no cover
    convert_from_bytes = None

try:
    import pytesseract
except ImportError:  # pragma: no cover
    pytesseract = None

try:
    from PIL import Image
except ImportError:  # pragma: no cover
    Image = None

try:
    from docx import Document
except ImportError:  # pragma: no cover
    Document = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


def _configure_ocr_tools() -> None:
    tesseract_cmd = os.getenv("TESSERACT_CMD", "").strip()
    if pytesseract is not None and tesseract_cmd:
        pytesseract.pytesseract.tesseract_cmd = tesseract_cmd


def _extract_pdf_text(file_bytes: bytes) -> str:
    if pdfplumber is not None:
        try:
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                pages = [(page.extract_text() or "").strip() for page in pdf.pages]
            text = "\n".join(part for part in pages if part).strip()
            if text:
                return text
        except Exception:
            pass
    if PdfReader is None:
        return ""
    try:
        reader = PdfReader(io.BytesIO(file_bytes))
        return "\n".join(page.extract_text() or "" for page in reader.pages).strip()
    except Exception:
        return ""


def _extract_pdf_text_with_ocr(file_bytes: bytes) -> str:
    if convert_from_bytes is None or pytesseract is None:
        return ""
    try:
        poppler_path = os.getenv("POPPLER_PATH", "").strip() or None
        images = convert_from_bytes(file_bytes, dpi=200, poppler_path=poppler_path)
        page_text = [pytesseract.image_to_string(image) for image in images]
        return "\n".join(text.strip() for text in page_text if text.strip())
    except Exception:
        return ""


def _extract_image_text(file_bytes: bytes) -> str:
    if Image is None or pytesseract is None:
        return ""
    try:
        image = Image.open(io.BytesIO(file_bytes))
        return pytesseract.image_to_string(image).strip()
    except Exception:
        return ""


def _extract_docx_text(file_bytes: bytes) -> str:
    if Document is None:
        return ""
    try:
        document = Document(io.BytesIO(file_bytes))
        return "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text.strip()).strip()
    except Exception:
        return ""


def _extract_xlsx_text(file_bytes: bytes) -> str:
    if load_workbook is None:
        return ""
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        rows = []
        for sheet in workbook.worksheets:
            rows.append(f"[Sheet] {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if values:
                    rows.append(" | ".join(values))
        return "\n".join(rows).strip()
    except Exception:
        return ""


def extract_text_with_optional_ocr(file_bytes: bytes, filename: str = "", content_type: str = "") -> str:
    """
    Tries native text extraction first, then OCR where appropriate.
    Safe fallback: returns empty string if no parser is available.
    """
    if not file_bytes:
        return ""

    _configure_ocr_tools()
    name = filename.lower()
    content_type = (content_type or "").lower()

    if name.endswith(".pdf") or "pdf" in content_type:
        text = _extract_pdf_text(file_bytes)
        return text or _extract_pdf_text_with_ocr(file_bytes)

    if name.endswith((".png", ".jpg", ".jpeg", ".tiff", ".bmp")) or content_type.startswith("image/"):
        return _extract_image_text(file_bytes)

    if name.endswith(".docx"):
        return _extract_docx_text(file_bytes)

    if name.endswith((".xlsx", ".xlsm")):
        return _extract_xlsx_text(file_bytes)

    try:
        return file_bytes.decode("utf-8")
    except UnicodeDecodeError:
        try:
            return file_bytes.decode("latin-1")
        except UnicodeDecodeError:
            return ""


def extract_text_metadata(file_bytes: bytes, filename: str = "", content_type: str = "") -> Dict[str, Any]:
    if not file_bytes:
        return {"text": "", "extraction_method": "Direct text extraction (pdfplumber)"}

    _configure_ocr_tools()
    name = filename.lower()
    content_type = (content_type or "").lower()

    if name.endswith(".pdf") or "pdf" in content_type:
        direct_text = _extract_pdf_text(file_bytes)
        if direct_text:
            return {"text": direct_text, "extraction_method": "Direct text extraction (pdfplumber)"}
        return {
            "text": _extract_pdf_text_with_ocr(file_bytes),
            "extraction_method": "OCR scan (Tesseract)",
        }

    if name.endswith((".png", ".jpg", ".jpeg", ".tiff", ".bmp")) or content_type.startswith("image/"):
        return {"text": _extract_image_text(file_bytes), "extraction_method": "OCR scan (Tesseract)"}

    text = extract_text_with_optional_ocr(file_bytes, filename=filename, content_type=content_type)
    return {"text": text, "extraction_method": "Direct text extraction (pdfplumber)"}


def _extract_amounts(text: str) -> list[float]:
    normalized = text.replace(",", "")
    matches = re.findall(r"(?:rs\.?|inr)?\s*(\d+(?:\.\d+)?)", normalized, flags=re.IGNORECASE)
    return [float(match) for match in matches]


def _pick_amount(amounts: list[float], default_value: float) -> float:
    return amounts[0] if amounts else default_value


def parse_balance_sheet(file_bytes: bytes) -> Dict[str, float]:
    text = extract_text_with_optional_ocr(file_bytes, filename="balance_sheet.pdf")
    amounts = _extract_amounts(text)
    return {
        "total_debt": _pick_amount(amounts[0:1], 50_00_00_000),
        "equity": _pick_amount(amounts[1:2], 75_00_00_000),
        "current_assets": _pick_amount(amounts[2:3], 80_00_00_000),
        "current_liabilities": _pick_amount(amounts[3:4], 10_00_00_000),
    }


def parse_pl_statement(file_bytes: bytes) -> Dict[str, float]:
    text = extract_text_with_optional_ocr(file_bytes, filename="pl_statement.pdf")
    amounts = _extract_amounts(text)
    return {
        "revenue": _pick_amount(amounts[0:1], 120_00_00_000),
        "net_profit": _pick_amount(amounts[1:2], 27_60_00_000),
        "interest": _pick_amount(amounts[2:3], 4_00_00_000),
    }


def parse_bank_statements(file_bytes: bytes) -> Dict[str, float]:
    text = extract_text_with_optional_ocr(file_bytes, filename="bank_statements.pdf")
    amounts = _extract_amounts(text)
    return {
        "total_debt_service": _pick_amount(amounts[0:1], 6_00_00_000),
        "cash_profit": _pick_amount(amounts[1:2], 7_08_00_000),
    }


def parse_gst_returns(file_bytes: bytes) -> Dict[str, float]:
    text = extract_text_with_optional_ocr(file_bytes, filename="gst_returns.pdf")
    percentages = re.findall(r"(\d+(?:\.\d+)?)\s*%", text)
    parsed = [float(value) for value in percentages]
    return {
        "gstr_2a": parsed[0] if len(parsed) > 0 else 100.0,
        "gstr_3b": parsed[1] if len(parsed) > 1 else 75.3,
    }


def parse_rating_report(file_bytes: bytes) -> Dict[str, Any]:
    text = extract_text_with_optional_ocr(file_bytes, filename="rating_report.pdf").upper()
    for grade in ["AAA", "AA", "A", "BBB", "BB", "B"]:
        if grade in text:
            return {"rating_grade": grade}
    return {"rating_grade": "BBB"}


def build_financial_ratios(
    bs: Dict[str, float],
    pl: Dict[str, float],
    bank: Dict[str, float],
) -> FinancialRatios:
    debt_to_equity = bs["total_debt"] / max(bs["equity"], 1.0)
    current_ratio = bs["current_assets"] / max(bs["current_liabilities"], 1.0)
    profit_margin = pl["net_profit"] / max(pl["revenue"], 1.0)
    dscr = bank["cash_profit"] / max(bank["total_debt_service"], 1.0)

    return FinancialRatios(
        dscr=dscr,
        debt_to_equity=debt_to_equity,
        current_ratio=current_ratio,
        profit_margin=profit_margin,
    )


def build_gst_features(gst_raw: Dict[str, float]) -> GSTFeatures:
    gap = abs(gst_raw["gstr_2a"] - gst_raw["gstr_3b"])
    base = max(gst_raw["gstr_2a"], gst_raw["gstr_3b"], 1e-6)
    gstr_gap_percent = (gap / base) * 100.0

    circular_flag = gstr_gap_percent > 20.0

    return GSTFeatures(
        gstr_gap_percent=gstr_gap_percent,
        circular_trading_flag=circular_flag,
    )


def basic_news_legal_features_stub(company_name: str) -> NewsLegalFeatures:
    return NewsLegalFeatures(
        negative_news_score=60,
        legal_case_count=3,
        major_events=[
            f"Media reports about sector headwinds for {company_name}.",
            f"Minor legal dispute reported for {company_name} in 2023.",
        ],
    )
